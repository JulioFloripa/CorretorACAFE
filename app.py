import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF, XPos, YPos
import tempfile
import zipfile
import os
import traceback
import base64
from PIL import Image
import requests
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Protection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.data_validation import DataValidation

# Configurar matplotlib para usar backend não-interativo
import matplotlib
matplotlib.use('Agg')

# --------------------------
# CONFIGURAÇÕES INICIAIS
# --------------------------
st.set_page_config(
    page_title="Corretor ACAFE Fleming", 
    layout="wide",
    page_icon="🎓",
    initial_sidebar_state="expanded"
)

# Inicializar estado da sessão
if 'processamento_concluido' not in st.session_state:
    st.session_state.processamento_concluido = False
if 'dados_processados' not in st.session_state:
    st.session_state.dados_processados = None
if 'logos_carregadas' not in st.session_state:
    st.session_state.logos_carregadas = False

# --------------------------
# FUNÇÕES PARA LOGOS
# --------------------------

@st.cache_data
def carregar_logos():
    """Carrega as logos do repositório GitHub"""
    logos = {}
    
    # URLs das logos no GitHub
    urls_logos = {
        'acafe': 'https://raw.githubusercontent.com/JulioFloripa/CorretorACAFE/main/logo-acafe.png',
        'fleming': 'https://raw.githubusercontent.com/JulioFloripa/CorretorACAFE/main/logo_fleming.png'
    }
    
    for nome, url in urls_logos.items():
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                # Salvar logo localmente
                logo_path = f"/tmp/logo_{nome}.png"
                with open(logo_path, 'wb') as f:
                    f.write(response.content)
                logos[nome] = logo_path
                
                # Converter para base64 para uso na interface
                logos[f'{nome}_b64'] = base64.b64encode(response.content).decode()
            else:
                logos[nome] = None
        except Exception as e:
            logos[nome] = None
    
    return logos

# Carregar logos
if not st.session_state.logos_carregadas:
    with st.spinner("🔄 Carregando logos oficiais..."):
        st.session_state.logos = carregar_logos()
        st.session_state.logos_carregadas = True

logos = st.session_state.logos

# --------------------------
# FUNÇÃO PARA CRIAR TEMPLATE EXCEL
# --------------------------

@st.cache_data
def criar_template_excel():
    """Cria template Excel com formatação ACAFE"""
    
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Cores tema ACAFE
    cor_verde_acafe = "2D5A3D"
    cor_verde_claro = "E8F5F3"
    cor_branco = "FFFFFF"
    
    # ===== ABA RESPOSTAS =====
    ws_respostas = wb.active
    ws_respostas.title = "RESPOSTAS"
    
    # Cabeçalhos da aba RESPOSTAS
    headers_respostas = ["ID", "Nome", "Sede"] + [f"Questão {i:02d}" for i in range(1, 71)]
    
    # Aplicar cabeçalhos
    for col, header in enumerate(headers_respostas, 1):
        cell = ws_respostas.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color=cor_verde_acafe, end_color=cor_verde_acafe, fill_type="solid")
        cell.font = Font(color=cor_branco, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados de exemplo
    exemplos_respostas = [
        [1, "João Silva Santos", "CRICIÚMA"] + ["A"] * 70,
        [2, "Maria Oliveira Costa", "TUBARÃO"] + ["B"] * 70,
        [3, "Pedro Souza Lima", "ARARANGUÁ"] + ["C"] * 70
    ]
    
    for row_idx, exemplo in enumerate(exemplos_respostas, 2):
        for col_idx, valor in enumerate(exemplo, 1):
            cell = ws_respostas.cell(row=row_idx, column=col_idx, value=valor)
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color=cor_verde_claro, end_color=cor_verde_claro, fill_type="solid")
    
    # Validação de dados para respostas (A, B, C, D, E)
    dv = DataValidation(type="list", formula1='"A,B,C,D,E"', allow_blank=True)
    dv.error = "Por favor, insira apenas A, B, C, D ou E"
    dv.errorTitle = "Entrada Inválida"
    ws_respostas.add_data_validation(dv)
    
    # Aplicar validação nas colunas de questões
    for col in range(4, 74):  # Colunas D até BU (questões 01-70)
        dv.add(f"{openpyxl.utils.get_column_letter(col)}2:{openpyxl.utils.get_column_letter(col)}1000")
    
    # Ajustar largura das colunas
    ws_respostas.column_dimensions['A'].width = 8   # ID
    ws_respostas.column_dimensions['B'].width = 25  # Nome
    ws_respostas.column_dimensions['C'].width = 15  # Sede
    for col in range(4, 74):
        ws_respostas.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 4
    
    # ===== ABA GABARITO =====
    ws_gabarito = wb.create_sheet("GABARITO")
    
    # Cabeçalhos da aba GABARITO
    headers_gabarito = ["Questão", "Resposta", "Disciplina"]
    
    for col, header in enumerate(headers_gabarito, 1):
        cell = ws_gabarito.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color=cor_verde_acafe, end_color=cor_verde_acafe, fill_type="solid")
        cell.font = Font(color=cor_branco, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Exemplo de gabarito
    disciplinas = ["Matemática", "Português", "História", "Geografia", "Biologia", "Física", "Química", "Inglês", "Espanhol"]
    exemplos_gabarito = []
    
    for i in range(1, 71):
        if i <= 56:
            disciplina = disciplinas[(i-1) % 7]  # Distribui entre as primeiras 7 disciplinas
        else:
            # Questões 57-70 são de línguas (Inglês e Espanhol)
            if i % 2 == 1:
                disciplina = "Inglês"
            else:
                disciplina = "Espanhol"
        
        resposta = ["A", "B", "C", "D", "E"][(i-1) % 5]
        exemplos_gabarito.append([i, resposta, disciplina])
    
    for row_idx, exemplo in enumerate(exemplos_gabarito, 2):
        for col_idx, valor in enumerate(exemplo, 1):
            cell = ws_gabarito.cell(row=row_idx, column=col_idx, value=valor)
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color=cor_verde_claro, end_color=cor_verde_claro, fill_type="solid")
    
    # Validação para respostas do gabarito
    dv_gabarito = DataValidation(type="list", formula1='"A,B,C,D,E"', allow_blank=False)
    dv_gabarito.error = "Por favor, insira apenas A, B, C, D ou E"
    dv_gabarito.errorTitle = "Entrada Inválida"
    ws_gabarito.add_data_validation(dv_gabarito)
    dv_gabarito.add("B2:B1000")
    
    # Ajustar largura das colunas
    ws_gabarito.column_dimensions['A'].width = 12  # Questão
    ws_gabarito.column_dimensions['B'].width = 12  # Resposta
    ws_gabarito.column_dimensions['C'].width = 20  # Disciplina
    
    # ===== ABA INSTRUÇÕES =====
    ws_instrucoes = wb.create_sheet("INSTRUÇÕES")
    
    instrucoes_texto = [
        ["TEMPLATE SIMULADO ACAFE - COLÉGIO FLEMING", ""],
        ["", ""],
        ["INSTRUÇÕES DE USO:", ""],
        ["", ""],
        ["1. ABA RESPOSTAS:", ""],
        ["   • Preencha o ID único de cada aluno", ""],
        ["   • Insira o nome completo do aluno", ""],
        ["   • Indique a sede (CRICIÚMA, TUBARÃO, etc.)", ""],
        ["   • Preencha as respostas nas colunas Questão 01 a 70", ""],
        ["   • Use apenas as letras: A, B, C, D, E", ""],
        ["", ""],
        ["2. ABA GABARITO:", ""],
        ["   • Questão: Número da questão (1 a 70)", ""],
        ["   • Resposta: Resposta correta (A, B, C, D, E)", ""],
        ["   • Disciplina: Nome da matéria", ""],
        ["", ""],
        ["3. QUESTÕES DE LÍNGUAS:", ""],
        ["   • Questões 57-70 podem ser Inglês OU Espanhol", ""],
        ["   • O sistema permite questões com mesmo número", ""],
        ["   • para disciplinas diferentes", ""],
        ["", ""],
        ["4. VALIDAÇÃO:", ""],
        ["   • Células têm validação automática", ""],
        ["   • Só aceita respostas válidas (A-E)", ""],
        ["   • Formatação tema ACAFE aplicada", ""],
        ["", ""],
        ["DESENVOLVIDO PARA COLÉGIO FLEMING", ""],
        ["Sistema de Correção ACAFE v4.0", ""]
    ]
    
    for row_idx, (texto, _) in enumerate(instrucoes_texto, 1):
        cell = ws_instrucoes.cell(row=row_idx, column=1, value=texto)
        if "TEMPLATE" in texto or "INSTRUÇÕES" in texto or "DESENVOLVIDO" in texto:
            cell.font = Font(bold=True, size=14, color=cor_verde_acafe)
        elif texto.startswith(("1.", "2.", "3.", "4.")):
            cell.font = Font(bold=True, color=cor_verde_acafe)
        else:
            cell.font = Font(color="333333")
    
    ws_instrucoes.column_dimensions['A'].width = 50
    
    # Proteger planilha (opcional - desabilitado para facilitar edição)
    # ws_respostas.protection = SheetProtection(password="acafe2024")
    # ws_gabarito.protection = SheetProtection(password="acafe2024")
    
    # Salvar em bytes
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()

# --------------------------
# CONFIGURAÇÕES DE ESTILO
# --------------------------

def load_css():
    """Carrega CSS customizado para tema verde ACAFE"""
    st.markdown("""
    <style>
    /* Tema principal verde ACAFE */
    .main {
        background: linear-gradient(135deg, #f8fffe 0%, #e8f5f3 100%);
    }
    
    /* Header customizado com logos */
    .header-acafe {
        background: linear-gradient(90deg, #2d5a3d 0%, #4a8c6a 100%);
        padding: 1.5rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(45, 90, 61, 0.2);
        display: flex;
        align-items: center;
        justify-content: space-between;
    }
    
    .header-content {
        text-align: center;
        flex-grow: 1;
    }
    
    .header-acafe h1 {
        color: white;
        font-size: 2.5rem;
        font-weight: bold;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .header-acafe p {
        color: #e8f5f3;
        font-size: 1.2rem;
        margin-top: 0.5rem;
        font-style: italic;
    }
    
    .logo-header {
        width: 80px;
        height: 80px;
        background: white;
        border-radius: 50%;
        padding: 10px;
        box-shadow: 0 4px 10px rgba(0,0,0,0.2);
    }
    
    .logo-header img {
        width: 100%;
        height: 100%;
        object-fit: contain;
        border-radius: 50%;
    }
    
    /* Botões customizados */
    .stButton > button {
        background: linear-gradient(45deg, #4a8c6a, #2d5a3d);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.5rem 1rem;
        font-weight: bold;
        transition: all 0.3s ease;
        box-shadow: 0 2px 5px rgba(45, 90, 61, 0.3);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 10px rgba(45, 90, 61, 0.4);
    }
    
    /* Métricas customizadas */
    .metric-container {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #4a8c6a;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        margin: 0.5rem 0;
    }
    
    /* Progress bar customizada */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #4a8c6a, #2d5a3d);
    }
    
    /* Alertas customizados */
    .stAlert {
        border-radius: 10px;
        border-left: 4px solid #4a8c6a;
    }
    
    /* Footer */
    .footer {
        text-align: center;
        padding: 2rem;
        color: #2d5a3d;
        font-style: italic;
        border-top: 2px solid #e8f5f3;
        margin-top: 3rem;
    }
    </style>
    """, unsafe_allow_html=True)

def show_header():
    """Mostra header customizado com logos oficiais"""
    # Preparar logos para o header
    logo_acafe_html = ""
    logo_fleming_html = ""
    
    if logos.get('acafe_b64'):
        logo_acafe_html = f'<div class="logo-header"><img src="data:image/png;base64,{logos["acafe_b64"]}" alt="ACAFE"></div>'
    
    if logos.get('fleming_b64'):
        logo_fleming_html = f'<div class="logo-header"><img src="data:image/png;base64,{logos["fleming_b64"]}" alt="Fleming"></div>'
    
    st.markdown(f"""
    <div class="header-acafe">
        {logo_acafe_html}
        <div class="header-content">
            <h1>Corretor ACAFE Fleming</h1>
            <p>Sistema Inteligente de Correção de Simulados</p>
        </div>
        {logo_fleming_html}
    </div>
    """, unsafe_allow_html=True)

# --------------------------
# FUNÇÕES DE VALIDAÇÃO
# --------------------------

def validar_arquivo_excel(dados):
    """Valida se o arquivo Excel tem a estrutura esperada"""
    erros = []
    
    # Verificar se as abas existem
    if "RESPOSTAS" not in dados:
        erros.append("❌ Aba 'RESPOSTAS' não encontrada no arquivo")
    if "GABARITO" not in dados:
        erros.append("❌ Aba 'GABARITO' não encontrada no arquivo")
    
    if erros:
        return False, erros
    
    respostas = dados["RESPOSTAS"]
    gabarito = dados["GABARITO"]
    
    # Verificar colunas obrigatórias na aba RESPOSTAS
    colunas_obrigatorias_respostas = ["ID", "Nome"]
    for col in colunas_obrigatorias_respostas:
        if col not in respostas.columns:
            erros.append(f"❌ Coluna '{col}' não encontrada na aba RESPOSTAS")
    
    # Verificar colunas obrigatórias na aba GABARITO
    colunas_obrigatorias_gabarito = ["Questão", "Resposta", "Disciplina"]
    for col in colunas_obrigatorias_gabarito:
        if col not in gabarito.columns:
            erros.append(f"❌ Coluna '{col}' não encontrada na aba GABARITO")
    
    # Verificar se há dados
    if len(respostas) == 0:
        erros.append("❌ Aba RESPOSTAS está vazia")
    if len(gabarito) == 0:
        erros.append("❌ Aba GABARITO está vazia")
    
    return len(erros) == 0, erros

def validar_dados_gabarito(gabarito):
    """Valida os dados do gabarito"""
    erros = []
    
    # Verificar questões duplicadas APENAS dentro da mesma disciplina
    for disciplina in gabarito['Disciplina'].unique():
        if pd.isna(disciplina):
            continue
        
        gabarito_disciplina = gabarito[gabarito['Disciplina'] == disciplina]
        questoes_duplicadas = gabarito_disciplina[gabarito_disciplina.duplicated(subset=['Questão'], keep=False)]
        
        if len(questoes_duplicadas) > 0:
            questoes_dup = questoes_duplicadas['Questão'].unique().tolist()
            erros.append(f"❌ Questões duplicadas em {disciplina}: {questoes_dup}")
    
    # Verificar se há valores nulos
    if gabarito['Questão'].isnull().any():
        erros.append("❌ Há questões com número vazio no gabarito")
    if gabarito['Resposta'].isnull().any():
        erros.append("❌ Há questões sem resposta no gabarito")
    if gabarito['Disciplina'].isnull().any():
        erros.append("❌ Há questões sem disciplina no gabarito")
    
    # Verificar questões de línguas estrangeiras (informativo)
    linguas = ['Inglês', 'Espanhol', 'Ingles', 'Espanol']
    questoes_linguas = gabarito[gabarito['Disciplina'].isin(linguas)]
    
    if len(questoes_linguas) > 0:
        st.info(f"ℹ️ Detectadas {len(questoes_linguas)} questões de línguas estrangeiras. Questões com mesmo número são permitidas para Inglês/Espanhol.")
    
    return len(erros) == 0, erros

# --------------------------
# FUNÇÕES AUXILIARES OTIMIZADAS
# --------------------------

@st.cache_data
def corrigir_respostas_otimizado(df_respostas, gabarito, mapa_disciplinas):
    """Corrige as respostas dos alunos baseado no gabarito - VERSÃO OTIMIZADA"""
    respostas = df_respostas.copy()
    
    # Criar dicionário de gabarito para acesso rápido
    gabarito_dict = {}
    for _, row in gabarito.iterrows():
        questao = int(row["Questão"])
        resposta_correta = str(row["Resposta"]).strip().upper()
        gabarito_dict[questao] = resposta_correta
    
    # Processar todas as questões de uma vez (vetorizado)
    for questao, resposta_correta in gabarito_dict.items():
        col = f"Questão {questao:02d}"
        col_ok = f"Q{questao}_OK"
        
        if col in respostas.columns:
            # Comparação vetorizada
            respostas[col_ok] = (
                respostas[col].astype(str).str.strip().str.upper() == resposta_correta
            )
        else:
            respostas[col_ok] = False
    
    return respostas

def resultados_disciplina_otimizado(linha, mapa_disciplinas):
    """Calcula os resultados por disciplina para um aluno - OTIMIZADO"""
    resultados = []
    for disc, questoes in mapa_disciplinas.items():
        # Usar list comprehension para melhor performance
        acertos = sum(linha.get(f"Q{int(q)}_OK", False) for q in questoes)
        total = len(questoes)
        perc = round(100 * acertos / total, 1) if total > 0 else 0
        resultados.append((disc, acertos, total, perc))
    return resultados

def gerar_graficos_otimizado(nome, posicao, percentual, df_boletim, media_df, ranking_df, pasta):
    """Gera os gráficos para o boletim individual - VERSÃO OTIMIZADA"""
    try:
        labels = df_boletim["Disciplina"].tolist()
        aluno_vals = df_boletim["%"].values
        media_vals = media_df["%"].values

        # Configurar cores tema ACAFE
        cor_principal = '#2d5a3d'
        cor_secundaria = '#4a8c6a'
        cor_destaque = '#6bb77b'
        
        # Configurar estilo dos gráficos uma vez
        plt.style.use('default')
        plt.rcParams.update({
            'font.size': 10,
            'axes.titlesize': 14,
            'axes.labelsize': 12,
            'xtick.labelsize': 10,
            'ytick.labelsize': 10,
            'legend.fontsize': 12
        })
        
        graficos_paths = []
        
        # Gráfico de Barras (mais importante)
        if len(labels) > 0:
            x = np.arange(len(labels))
            bar_width = 0.35
            fig, ax = plt.subplots(figsize=(14, 8))
            
            bars1 = ax.bar(x - bar_width/2, aluno_vals, bar_width, label=nome, 
                          color=cor_principal, alpha=0.8, edgecolor='white', linewidth=1)
            bars2 = ax.bar(x + bar_width/2, media_vals, bar_width, label="Média Turma", 
                          color=cor_secundaria, alpha=0.7, edgecolor='white', linewidth=1)
            
            # Adicionar valores nas barras
            for i, v in enumerate(aluno_vals):
                ax.text(i - bar_width/2, v + 1.5, f"{v:.1f}%", ha="center", fontsize=10, 
                       fontweight='bold', color=cor_principal)
            for i, v in enumerate(media_vals):
                ax.text(i + bar_width/2, v + 1.5, f"{v:.1f}%", ha="center", fontsize=10, 
                       color=cor_secundaria)
                
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=11)
            ax.set_ylabel("Percentual de Acertos (%)", fontsize=12, fontweight='bold')
            ax.set_title(f"Desempenho por Disciplina - {nome}", fontsize=16, fontweight='bold', 
                        color=cor_principal, pad=20)
            ax.legend(fontsize=12)
            ax.grid(axis='y', alpha=0.3)
            ax.set_ylim(0, 105)
            
            barras_path = os.path.join(pasta, f"{nome}_barras.png")
            plt.savefig(barras_path, bbox_inches="tight", dpi=150, facecolor='white')
            plt.close()
            graficos_paths.append(barras_path)
        else:
            graficos_paths.append(None)

        # Gráfico Radar (se houver disciplinas suficientes)
        if len(labels) >= 3:
            angles = np.linspace(0, 2 * np.pi, len(labels), endpoint=False).tolist()
            aluno_circ = np.concatenate((aluno_vals, [aluno_vals[0]]))
            media_circ = np.concatenate((media_vals, [media_vals[0]]))
            angles += [angles[0]]

            fig = plt.figure(figsize=(8, 8))
            ax = plt.subplot(111, polar=True)
            ax.plot(angles, aluno_circ, "o-", label=nome, linewidth=3, color=cor_principal, markersize=8)
            ax.fill(angles, aluno_circ, alpha=0.3, color=cor_principal)
            ax.plot(angles, media_circ, "s--", label="Média da Turma", color=cor_secundaria, linewidth=2, markersize=6)
            ax.fill(angles, media_circ, alpha=0.1, color=cor_secundaria)
            ax.set_thetagrids(np.degrees(angles[:-1]), labels, fontsize=10)
            ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1), fontsize=12)
            ax.set_ylim(0, 100)
            ax.grid(True, alpha=0.3)
            plt.title(f"Desempenho Radar - {nome}", fontsize=14, fontweight='bold', color=cor_principal, pad=20)
            radar_path = os.path.join(pasta, f"{nome}_radar.png")
            plt.savefig(radar_path, bbox_inches="tight", dpi=150, facecolor='white')
            plt.close()
            graficos_paths.append(radar_path)
        else:
            graficos_paths.append(None)

        # Distribuição das notas
        fig, ax = plt.subplots(figsize=(12, 7))
        n, bins, patches = ax.hist(ranking_df["Percentual"]*100, bins=min(12, len(ranking_df)), 
                                  color=cor_destaque, edgecolor=cor_principal, alpha=0.7, linewidth=1.5)
        
        # Colorir a barra onde o aluno está
        for i, patch in enumerate(patches):
            if bins[i] <= percentual <= bins[i+1]:
                patch.set_color(cor_principal)
                patch.set_alpha(0.9)
        
        ax.axvline(percentual, color='red', linewidth=4, 
                  label=f"{nome} ({percentual:.1f}%)", linestyle='--', alpha=0.8)
        ax.set_xlabel("Percentual de Acertos (%)", fontsize=12, fontweight='bold')
        ax.set_ylabel("Número de Estudantes", fontsize=12, fontweight='bold')
        ax.set_title("Distribuição das Notas da Turma", fontsize=16, fontweight='bold', 
                    color=cor_principal, pad=20)
        ax.legend(fontsize=12)
        ax.grid(alpha=0.3)
        
        dist_path = os.path.join(pasta, f"{nome}_dist.png")
        plt.savefig(dist_path, bbox_inches="tight", dpi=150, facecolor='white')
        plt.close()
        graficos_paths.append(dist_path)

        # Ranking
        fig, ax = plt.subplots(figsize=(12, 7))
        ax.plot(ranking_df["Posição"], ranking_df["Percentual"]*100, "o-", 
               color=cor_secundaria, markersize=8, linewidth=3, alpha=0.7, label="Outros alunos")
        ax.scatter(posicao, percentual, color='red', s=200, 
                  label=f"{nome} - {posicao}º lugar", zorder=5, edgecolor='darkred', linewidth=2)
        
        # Destacar top 3
        top3 = ranking_df.head(3)
        ax.scatter(top3["Posição"], top3["Percentual"]*100, color='gold', s=150, 
                  zorder=4, edgecolor='orange', linewidth=2, alpha=0.8, label="Top 3")
        
        ax.set_xlabel("Posição no Ranking", fontsize=12, fontweight='bold')
        ax.set_ylabel("Percentual de Acertos (%)", fontsize=12, fontweight='bold')
        ax.set_title("Ranking da Turma", fontsize=16, fontweight='bold', color=cor_principal, pad=20)
        ax.legend(fontsize=12)
        ax.grid(alpha=0.3)
        
        rank_path = os.path.join(pasta, f"{nome}_rank.png")
        plt.savefig(rank_path, bbox_inches="tight", dpi=150, facecolor='white')
        plt.close()
        graficos_paths.append(rank_path)

        return graficos_paths
    
    except Exception as e:
        st.error(f"Erro ao gerar gráficos para {nome}: {str(e)}")
        return [None, None, None, None]

class BoletimPDF(FPDF):
    def __init__(self):
        super().__init__()
        self.logo_acafe_path = logos.get('acafe')
        self.logo_fleming_path = logos.get('fleming')
    
    def header(self):
        """Header melhorado com logos oficiais - SEM WARNINGS"""
        # Fundo verde no header
        self.set_fill_color(45, 90, 61)  # Verde ACAFE
        self.rect(0, 0, 210, 45, 'F')
        
        # Logo ACAFE (esquerda)
        if self.logo_acafe_path and os.path.exists(self.logo_acafe_path):
            try:
                self.image(self.logo_acafe_path, 15, 8, 30)
            except Exception:
                pass
        
        # Logo Fleming (direita)
        if self.logo_fleming_path and os.path.exists(self.logo_fleming_path):
            try:
                self.image(self.logo_fleming_path, 165, 8, 30)
            except Exception:
                pass
        
        # Título central
        self.set_font("Helvetica", "B", 20)
        self.set_text_color(255, 255, 255)  # Branco
        self.set_y(15)
        self.cell(0, 8, "SIMULADO ACAFE", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        
        self.set_font("Helvetica", "B", 16)
        self.cell(0, 8, "COLEGIO FLEMING", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        
        self.set_font("Helvetica", "", 12)
        self.cell(0, 6, "Relatorio Individual de Desempenho", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        
        # Linha decorativa
        self.set_draw_color(255, 255, 255)
        self.set_line_width(1)
        self.line(20, 42, 190, 42)
        
        self.set_text_color(0, 0, 0)  # Voltar para preto
        self.ln(18)

    def add_aluno_info(self, nome, posicao, percentual, media_turma, aluno_data=None):
        """Informações do aluno com design melhorado - SEM WARNINGS"""
        # Caixa principal
        self.set_fill_color(240, 248, 245)  # Verde muito claro
        self.set_draw_color(45, 90, 61)  # Verde escuro
        self.set_line_width(1)
        self.rect(10, self.get_y(), 190, 50, 'DF')
        
        # Título da seção
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(45, 90, 61)
        self.set_y(self.get_y() + 8)
        self.cell(0, 8, "INFORMACOES DO ESTUDANTE", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        
        # Informações em duas colunas
        y_start = self.get_y() + 3
        
        # Coluna esquerda
        self.set_font("Helvetica", "B", 12)
        self.set_text_color(0, 0, 0)
        self.set_y(y_start)
        self.set_x(15)
        self.cell(90, 7, f"Nome: {nome}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.set_x(15)
        self.cell(90, 7, f"Posicao no Ranking: {posicao}º lugar", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        if aluno_data and 'Sede' in aluno_data:
            self.set_x(15)
            self.cell(90, 7, f"Sede: {aluno_data['Sede']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        # Coluna direita
        self.set_y(y_start)
        self.set_x(110)
        self.cell(90, 7, f"Nota Individual: {percentual:.1f}%", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.set_x(110)
        self.cell(90, 7, f"Media da Turma: {media_turma:.1f}%", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        # Diferença com cor
        diferenca = percentual - media_turma
        self.set_x(110)
        if diferenca > 0:
            self.set_text_color(0, 128, 0)  # Verde
            self.cell(90, 7, f"Diferenca: +{diferenca:.1f}% (acima)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        else:
            self.set_text_color(255, 0, 0)  # Vermelho
            self.cell(90, 7, f"Diferenca: {diferenca:.1f}% (abaixo)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.set_text_color(0, 0, 0)  # Voltar para preto
        self.ln(18)

    def add_table(self, df):
        """Tabela melhorada com cores alternadas - SEM WARNINGS"""
        # Título da tabela
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(45, 90, 61)
        self.cell(0, 10, "DESEMPENHO POR DISCIPLINA", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.ln(5)
        
        # Cabeçalho da tabela
        self.set_fill_color(45, 90, 61)  # Verde ACAFE
        self.set_text_color(255, 255, 255)  # Branco
        self.set_font("Helvetica", "B", 10)
        
        self.cell(50, 10, "Disciplina", 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        self.cell(25, 10, "Acertos", 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        self.cell(25, 10, "Total", 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        self.cell(30, 10, "Nota (%)", 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        self.cell(30, 10, "Media (%)", 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        self.cell(30, 10, "Diferenca", 1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True)
        
        # Dados da tabela
        self.set_font("Helvetica", "", 9)
        
        for i, (_, row) in enumerate(df.iterrows()):
            # Alternar cores das linhas
            if i % 2 == 0:
                self.set_fill_color(248, 255, 254)  # Verde muito claro
            else:
                self.set_fill_color(255, 255, 255)  # Branco
            
            self.set_text_color(0, 0, 0)
            disciplina = str(row["Disciplina"])[:22]  # Limitar tamanho
            self.cell(50, 8, disciplina, 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='L', fill=True)
            self.cell(25, 8, str(row["Acertos"]), 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
            self.cell(25, 8, str(row["Total"]), 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
            self.cell(30, 8, f"{row['%']:.1f}%", 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
            self.cell(30, 8, f"{row['Media Turma']:.1f}%", 1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
            
            diferenca = row['Diferenca']
            texto_dif = f"+{diferenca:.1f}%" if diferenca > 0 else f"{diferenca:.1f}%"
            self.cell(30, 8, texto_dif, 1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True)
        
        self.set_text_color(0, 0, 0)  # Voltar para preto
        self.ln(12)

    def add_image(self, path, largura=180, titulo=""):
        """Adiciona imagem com título - SEM WARNINGS"""
        if path and os.path.exists(path):
            try:
                if titulo:
                    self.set_font("Helvetica", "B", 12)
                    self.set_text_color(45, 90, 61)
                    self.cell(0, 10, titulo, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
                    self.ln(3)
                
                x_pos = (210 - largura) / 2
                self.image(path, x=x_pos, w=largura)
                self.ln(12)
                
            except Exception as e:
                self.set_font("Helvetica", "", 10)
                self.set_text_color(255, 0, 0)
                self.cell(0, 10, f"Erro ao carregar grafico: {str(e)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
                self.set_text_color(0, 0, 0)

    def footer(self):
        """Footer melhorado - SEM WARNINGS"""
        self.set_y(-25)
        
        # Linha decorativa
        self.set_draw_color(45, 90, 61)
        self.set_line_width(0.8)
        self.line(20, self.get_y(), 190, self.get_y())
        
        self.set_font("Helvetica", "", 9)
        self.set_text_color(100, 100, 100)
        self.ln(5)
        self.cell(0, 5, f"Pagina {self.page_no()}", new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        self.ln(4)
        self.cell(0, 5, "Sistema de Correcao ACAFE - Colegio Fleming | v4.0", new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')

# --------------------------
# APLICAR CSS E HEADER
# --------------------------
load_css()
show_header()

# --------------------------
# SIDEBAR
# --------------------------

with st.sidebar:
    st.markdown("### 🎓 **Instruções ACAFE**")
    
    # Mostrar status das logos
    if logos.get('acafe') and logos.get('fleming'):
        st.success("✅ Logos oficiais carregadas!")
    else:
        st.warning("⚠️ Algumas logos não foram carregadas")
    
    # BOTÃO PARA BAIXAR TEMPLATE
    st.markdown("### 📋 **Template Excel**")
    
    template_excel = criar_template_excel()
    st.download_button(
        label="📥 **Baixar Template Excel**",
        data=template_excel,
        file_name="Template_Simulado_ACAFE_Fleming.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Template pré-formatado com validação de dados e tema ACAFE",
        use_container_width=True
    )
    
    st.info("💡 **Use este template** para garantir que seu arquivo tenha a estrutura correta!")
    
    with st.expander("📊 **Aba RESPOSTAS**", expanded=False):
        st.markdown("""
        - **ID**: Número único do aluno
        - **Nome**: Nome completo
        - **Sede**: Unidade do colégio
        - **Questão 01-70**: Respostas (A, B, C, D, E)
        """)
    
    with st.expander("📝 **Aba GABARITO**", expanded=False):
        st.markdown("""
        - **Questão**: Número da questão (1-70)
        - **Resposta**: Resposta correta (A-E)
        - **Disciplina**: Nome da matéria
        """)
    
    st.markdown("### 📊 **Estatísticas**")
    if 'stats' in st.session_state:
        stats = st.session_state.stats
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("👥 Alunos", stats.get('total_alunos', 0))
            st.metric("📚 Disciplinas", stats.get('total_disciplinas', 0))
        with col2:
            st.metric("❓ Questões", stats.get('total_questoes', 0))
            if 'media_geral' in stats:
                st.metric("📈 Média", f"{stats['media_geral']:.1f}%")

# --------------------------
# INTERFACE PRINCIPAL
# --------------------------

# Botão de reset se processamento foi concluído
if st.session_state.processamento_concluido:
    st.success("✅ **Processamento concluído com sucesso!**")
    
    if st.button("🔄 **Processar Novo Arquivo**", type="primary"):
        st.session_state.processamento_concluido = False
        st.session_state.dados_processados = None
        st.rerun()
    
    # Mostrar dados processados se existirem
    if st.session_state.dados_processados:
        dados_proc = st.session_state.dados_processados
        
        # Mostrar ranking
        st.markdown("### 🏆 **Ranking da Turma**")
        
        # Top 3 destacado
        col1, col2, col3 = st.columns(3)
        top3 = dados_proc['ranking_df'].head(3)
        
        if len(top3) >= 1:
            with col1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #FFD700, #FFA500); padding: 1rem; border-radius: 15px; text-align: center; color: white; box-shadow: 0 4px 10px rgba(255,215,0,0.3);">
                    <h2 style="margin: 0;">🥇</h2>
                    <h4 style="margin: 0.5rem 0;">{top3.iloc[0]['Nome']}</h4>
                    <h3 style="margin: 0;">{top3.iloc[0]['Nota (%)']}%</h3>
                </div>
                """, unsafe_allow_html=True)
        
        if len(top3) >= 2:
            with col2:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #C0C0C0, #A0A0A0); padding: 1rem; border-radius: 15px; text-align: center; color: white; box-shadow: 0 4px 10px rgba(192,192,192,0.3);">
                    <h2 style="margin: 0;">🥈</h2>
                    <h4 style="margin: 0.5rem 0;">{top3.iloc[1]['Nome']}</h4>
                    <h3 style="margin: 0;">{top3.iloc[1]['Nota (%)']}%</h3>
                </div>
                """, unsafe_allow_html=True)
        
        if len(top3) >= 3:
            with col3:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #CD7F32, #B8860B); padding: 1rem; border-radius: 15px; text-align: center; color: white; box-shadow: 0 4px 10px rgba(205,127,50,0.3);">
                    <h2 style="margin: 0;">🥉</h2>
                    <h4 style="margin: 0.5rem 0;">{top3.iloc[2]['Nome']}</h4>
                    <h3 style="margin: 0;">{top3.iloc[2]['Nota (%)']}%</h3>
                </div>
                """, unsafe_allow_html=True)
        
        # Tabela do ranking
        st.dataframe(
            dados_proc['ranking_df'][["Posição", "Nome", "Nota (%)"]].head(10), 
            use_container_width=True,
            hide_index=True
        )
        
        # Médias por disciplina
        st.markdown("### 📊 **Médias por Disciplina**")
        st.dataframe(dados_proc['media_df'], use_container_width=True, hide_index=True)

else:
    # Interface de upload
    st.markdown("### 📚 Faça upload da planilha com as abas **RESPOSTAS** e **GABARITO**")
    
    arquivo = st.file_uploader(
        "📎 **Selecione o arquivo Excel**", 
        type=["xlsx"], 
        help="Arquivo deve conter as abas 'RESPOSTAS' e 'GABARITO'. Use o template acima para garantir compatibilidade!",
        key="file_uploader"
    )

    if arquivo:
        try:
            # Mostrar progresso
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            status_text.success("📖 Lendo arquivo Excel...")
            progress_bar.progress(10)
            
            # Ler arquivo Excel com engine otimizado
            dados = pd.read_excel(arquivo, sheet_name=None, engine='openpyxl')
            
            status_text.success("✅ Validando estrutura do arquivo...")
            progress_bar.progress(20)
            
            # Validar arquivo
            valido, erros = validar_arquivo_excel(dados)
            if not valido:
                st.error("**🚨 Problemas encontrados no arquivo:**")
                for erro in erros:
                    st.error(erro)
                st.stop()
            
            respostas = dados["RESPOSTAS"]
            gabarito = dados["GABARITO"]
            
            # Validar gabarito
            gabarito_valido, erros_gabarito = validar_dados_gabarito(gabarito)
            if not gabarito_valido:
                st.error("**🚨 Problemas encontrados no gabarito:**")
                for erro in erros_gabarito:
                    st.error(erro)
                st.stop()
            
            status_text.success("📊 Processando dados...")
            progress_bar.progress(30)
            
            # Estatísticas
            total_alunos = len(respostas)
            total_questoes = len(gabarito)
            disciplinas = gabarito['Disciplina'].unique()
            total_disciplinas = len(disciplinas)
            
            # Mostrar estatísticas principais
            st.markdown("### 📊 **Estatísticas do Simulado**")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown(f"""
                <div class="metric-container">
                    <h3 style="color: #2d5a3d; margin: 0;">👥 {total_alunos}</h3>
                    <p style="margin: 0; color: #666;">Alunos</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-container">
                    <h3 style="color: #2d5a3d; margin: 0;">❓ {total_questoes}</h3>
                    <p style="margin: 0; color: #666;">Questões</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="metric-container">
                    <h3 style="color: #2d5a3d; margin: 0;">📚 {total_disciplinas}</h3>
                    <p style="margin: 0; color: #666;">Disciplinas</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Processar dados
            status_text.success("🔄 Corrigindo respostas...")
            progress_bar.progress(40)
            
            # Mapeamento disciplinas
            mapa_disciplinas = {}
            for disciplina in gabarito['Disciplina'].unique():
                if pd.isna(disciplina):
                    continue
                questoes = gabarito[gabarito['Disciplina'] == disciplina]['Questão'].tolist()
                mapa_disciplinas[disciplina] = questoes

            # Usar função otimizada
            respostas_corr = corrigir_respostas_otimizado(respostas, gabarito, mapa_disciplinas)
            
            status_text.success("📈 Calculando ranking...")
            progress_bar.progress(50)

            # Ranking otimizado
            questoes_unicas = gabarito['Questão'].unique()
            
            # Calcular percentuais de forma vetorizada
            colunas_ok = [f"Q{int(q)}_OK" for q in questoes_unicas]
            colunas_ok_existentes = [col for col in colunas_ok if col in respostas_corr.columns]
            
            respostas_corr["Percentual"] = respostas_corr[colunas_ok_existentes].sum(axis=1) / len(questoes_unicas)

            ranking_df = respostas_corr[["ID", "Nome", "Percentual"]].sort_values("Percentual", ascending=False).reset_index(drop=True)
            ranking_df["Posição"] = ranking_df.index + 1
            ranking_df["Nota (%)"] = (ranking_df["Percentual"] * 100).round(1)
            media_turma = ranking_df["Percentual"].mean() * 100
            
            # Atualizar estatísticas
            with col4:
                st.markdown(f"""
                <div class="metric-container">
                    <h3 style="color: #2d5a3d; margin: 0;">📈 {media_turma:.1f}%</h3>
                    <p style="margin: 0; color: #666;">Média Geral</p>
                </div>
                """, unsafe_allow_html=True)

            # Salvar estatísticas
            st.session_state.stats = {
                'total_alunos': total_alunos,
                'total_questoes': total_questoes,
                'total_disciplinas': total_disciplinas,
                'media_geral': media_turma
            }
            
            status_text.success("📊 Calculando médias por disciplina...")
            progress_bar.progress(60)

            # Médias por disciplina otimizadas
            media_disciplinas = []
            for disc, questoes in mapa_disciplinas.items():
                colunas_disc = [f"Q{int(q)}_OK" for q in questoes if f"Q{int(q)}_OK" in respostas_corr.columns]
                if colunas_disc:
                    media_disc = respostas_corr[colunas_disc].mean().mean() * 100
                    media_disciplinas.append((disc, round(media_disc, 1)))
            
            media_df = pd.DataFrame(media_disciplinas, columns=["Disciplina", "%"])
            
            status_text.success("📄 Gerando boletins individuais...")
            progress_bar.progress(70)

            # Gerar boletins
            with tempfile.TemporaryDirectory() as tmpdir:
                zip_path = os.path.join(tmpdir, "boletins.zip")
                
                with zipfile.ZipFile(zip_path, "w") as zipf:
                    total_alunos = len(respostas_corr)
                    
                    for i, aluno in respostas_corr.iterrows():
                        # Atualizar progresso
                        progresso = 70 + (i / total_alunos) * 25
                        progress_bar.progress(int(progresso))
                        status_text.success(f"📄 Gerando boletim: {aluno['Nome']} ({i+1}/{total_alunos})")
                        
                        nome = aluno["Nome"].replace(" ", "_").replace("/", "_")
                        posicao = int(ranking_df.loc[ranking_df["ID"] == aluno["ID"], "Posição"].iloc[0])
                        percentual = aluno["Percentual"] * 100

                        resultados = resultados_disciplina_otimizado(aluno, mapa_disciplinas)
                        df_boletim = pd.DataFrame(resultados, columns=["Disciplina", "Acertos", "Total", "%"])
                        df_boletim["Media Turma"] = media_df["%"]
                        df_boletim["Diferenca"] = (df_boletim["%"] - media_df["%"]).round(1)

                        # Gráficos otimizados
                        graficos = gerar_graficos_otimizado(nome, posicao, percentual, df_boletim, media_df, ranking_df, tmpdir)

                        # PDF COM LOGOS OFICIAIS E SEM WARNINGS
                        try:
                            pdf = BoletimPDF()
                            pdf.add_page()
                            
                            # Informações do aluno
                            aluno_data = {'Sede': aluno.get('Sede', 'N/A')}
                            pdf.add_aluno_info(aluno["Nome"], posicao, percentual, media_turma, aluno_data)
                            
                            # Tabela
                            pdf.add_table(df_boletim)
                            
                            # Gráficos
                            titulos = [
                                "DESEMPENHO POR DISCIPLINA",
                                "GRAFICO RADAR - COMPARACAO COM A TURMA", 
                                "DISTRIBUICAO DAS NOTAS DA TURMA",
                                "POSICAO NO RANKING GERAL"
                            ]
                            
                            for grafico, titulo in zip(graficos, titulos):
                                if grafico:
                                    pdf.add_image(grafico, titulo=titulo)

                            pdf_path = os.path.join(tmpdir, f"Boletim_{nome}.pdf")
                            pdf.output(pdf_path)
                            zipf.write(pdf_path, f"Boletim_{nome}.pdf")
                        
                        except Exception as e:
                            st.warning(f"⚠️ Erro ao gerar PDF para {aluno['Nome']}: {str(e)}")
                            continue

                status_text.success("✅ Processamento concluído!")
                progress_bar.progress(100)
                
                # Salvar dados processados
                st.session_state.dados_processados = {
                    'ranking_df': ranking_df,
                    'media_df': media_df
                }
                
                # Botão de download
                with open(zip_path, "rb") as f:
                    st.markdown("### 🎉 **Boletins Prontos!**")
                    st.download_button(
                        "📥 **Baixar Todos os Boletins (ZIP)**", 
                        f.read(), 
                        "boletins_acafe_fleming.zip", 
                        "application/zip",
                        help=f"Arquivo contém {total_alunos} boletins individuais em PDF com logos oficiais",
                        use_container_width=True
                    )
                
                # Marcar como concluído
                st.session_state.processamento_concluido = True
                st.balloons()
                st.success(f"🎊 **{total_alunos} boletins gerados com sucesso!**")
                
        except Exception as e:
            st.error(f"❌ **Erro durante o processamento:** {str(e)}")
            with st.expander("🔍 **Detalhes técnicos do erro**"):
                st.code(traceback.format_exc())

# Footer
st.markdown("""
<div class="footer">
    <p><strong>Corretor ACAFE - Colégio Fleming</strong></p>
    <p>Desenvolvido com ❤️ para facilitar a correção de simulados</p>
    <p style="font-size: 0.8rem; opacity: 0.7;">Versão 4.0 FINAL - Template Excel | Performance Otimizada | Logos Oficiais | Sem Warnings</p>
</div>
""", unsafe_allow_html=True)
